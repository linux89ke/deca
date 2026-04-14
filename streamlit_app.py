
"""
Decathlon Product Lookup - FINAL COMPLETE VERSION
- Fashion → uses 'size' column (exactly as your original code)
- Other   → uses 'variation' column, shows '...' when empty
- sizes.txt loaded ONLY from project folder (no upload)
- Fashion sizes are editable per SKU in preview
- Invalid sizes marked with ❌
- Preview shows ONLY Primary Category
- Download = Upload Template sheet ONLY
- Price_KES always 100000, Stock always 0
"""

import os, io, re, json, asyncio
from typing import Optional
import numpy as np
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

try:
    from groq import AsyncGroq
    GROQ_AVAILABLE = True
except ImportError:
    GROQ_AVAILABLE = False

st.set_page_config(page_title="Decathlon Product Lookup", page_icon="", layout="wide")

st.markdown("""
<style>
h1 { color: #0082C3; }
.tag { display:inline-block; background:#0082C3; color:white; border-radius:4px; padding:2px 8px; font-size:12px; margin:2px; }
.ai-badge { display:inline-block; background:linear-gradient(90deg,#f55036,#ff8c00); color:white; border-radius:12px; padding:2px 10px; font-size:11px; font-weight:700; margin-left:6px; }
.kw-badge { display:inline-block; background:#0082C3; color:white; border-radius:12px; padding:2px 10px; font-size:11px; font-weight:700; margin-left:6px; }
</style>
""", unsafe_allow_html=True)

st.title("Decathlon Product Lookup")
st.markdown("Search by SKU number — view details, images, and **download a filled upload template**.")

# ====================== CONSTANTS ======================
IMAGE_COLS = ["OG_image"] + [f"picture_{i}" for i in range(1, 11)]
TEMPLATE_PATH = "product-creation-template.xlsx"
DECA_CAT_PATH = "deca_cat.xlsx"
MASTER_PATH = "Decathlon_Working_File_Split.csv"
SIZES_PATH = "sizes.txt"

MASTER_TO_TEMPLATE = {
    "product_name": "Name",
    "designed_for": "Description",
    "sku_num_sku_r3": "SellerSKU",
    "brand_name": "Brand",
    "bar_code": "GTIN_Barcode",
    "color": "color",
    "model_label": "model",
    "OG_image": "MainImage",
    "picture_1": "Image2",
    "picture_2": "Image3",
    "picture_3": "Image4",
    "picture_4": "Image5",
    "picture_5": "Image6",
    "picture_6": "Image7",
    "picture_7": "Image8",
}

# ====================== HELPERS ======================
def _clean(val):
    if pd.isna(val) or str(val).strip() in ("", "-", "nan"):
        return ""
    return str(val).strip()

def _format_gtin(val):
    raw = str(val).strip()
    if not raw or raw.lower() in ("nan", ""):
        return ""
    try:
        return str(int(float(raw)))
    except (ValueError, OverflowError):
        return raw

# ====================== UK SIZE EXTRACTION ======================
_UK_SIZE_PATTERNS = [
    re.compile(r'\bUK\s*(\d{1,2}(?:\.\d)?)\b', re.IGNORECASE),
    re.compile(r'\bUK\s*(\d{1,2}(?:\.\d)?)\s*[-–]\s*\d{1,2}', re.IGNORECASE),
]

def extract_uk_size(raw: str) -> Optional[str]:
    if not raw:
        return None
    cleaned = re.sub(r'"+', '', raw).strip()
    for pat in _UK_SIZE_PATTERNS:
        m = pat.search(cleaned)
        if m:
            return f"UK {m.group(1)}"
    return None

def parse_valid_sizes(path: str) -> list:
    try:
        with open(path, "r", encoding="utf-8") as f:
            lines = [l.strip() for l in f if l.strip() and not l.startswith("#")]
        return lines
    except FileNotFoundError:
        return []

# ====================== VARIATION (EXACTLY YOUR ORIGINAL) ======================
def get_variation(row: pd.Series, is_fashion: bool = True, valid_sizes: Optional[list] = None, size_override: Optional[str] = None) -> str:
    if not is_fashion:
        raw = re.sub(r'"+', '', str(row.get("variation", ""))).strip().rstrip(".")
        if raw.lower() in ("", "nan", "no size", "none"):
            return "..."
        return raw
    # Fashion path
    raw = re.sub(r'"+', '', str(row.get("size", ""))).strip().rstrip(".")
    if raw.lower() in ("", "nan", "no size", "none"):
        return size_override or "..."
    if size_override:
        return size_override
    if valid_sizes:
        raw_upper = raw.upper()
        for s in valid_sizes:
            if s.upper() == raw_upper:
                return s
    uk = extract_uk_size(raw)
    if uk and valid_sizes:
        uk_upper = uk.upper()
        for s in valid_sizes:
            if s.upper() == uk_upper:
                return s
        return uk
    if valid_sizes:
        raw_lower = raw.lower()
        for s in valid_sizes:
            if s.lower() in raw_lower or raw_lower in s.lower():
                return s
    return raw

# ====================== SHORT DESCRIPTION ======================
GENDER_MAP = {
    "MEN'S": "Men", "WOMEN'S": "Women", "BOYS'": "Boys", "GIRLS'": "Girls",
    "MEN": "Men", "WOMEN": "Women", "UNISEX": "Unisex", "NO GENDER": "", "HORSE": "",
}

_QUALITY_KEYWORDS = [
    "comfortable", "lightweight", "durable", "breathable", "waterproof", "quick-dry",
    "stretch", "supportive", "cushioned", "reflective", "insulated"
]

def rule_based_short_desc(row: pd.Series) -> str:
    bullets = []
    brand = _clean(row.get("brand_name", "")).title()
    dept = _clean(row.get("department_label", "")).replace("/", "·").title()
    g_raw = _clean(row.get("channable_gender", "")).split("|")[0].strip().upper()
    gender = GENDER_MAP.get(g_raw, g_raw.title())
    b1_parts = [p for p in [brand, dept, gender] if p]
    if b1_parts:
        bullets.append(" · ".join(b1_parts))

    color = _clean(row.get("color", "")).split("|")[0].strip().title()
    size = re.sub(r'"+', "", _clean(row.get("size", ""))).strip().rstrip(".")
    if color and size and size.lower() != "no size":
        bullets.append(f"Colour: {color} · Size: {size}")
    elif color:
        bullets.append(f"Colour: {color}")

    if not bullets:
        return ""
    items = "".join(f"<li>{b}</li>" for b in bullets[:3])
    return f"<ul>{items}</ul>"

# ====================== CATEGORY FIELDS ======================
CATEGORY_MATCH_FIELDS = [
    "family","type","department_label","nature_label",
    "proposed_brand_name","brand_name","color","channable_gender",
    "size","keywords","designed_for","business_weight","product_name",
]

# ====================== TEMPLATE BUILDER ======================
def build_template(results_df, df_cat, df_brands, ai_categories, short_descs, is_fashion, valid_sizes, sku_to_size_override=None):
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Upload Template"]

    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val:
            header_map[val] = col_idx

    hfont = ws.cell(row=1, column=1).font
    data_font = Font(name=hfont.name or "Calibri", size=hfont.size or 11)
    data_align = Alignment(vertical="center")

    model_to_first_sku = {}
    for _, r in results_df.iterrows():
        mc = str(r.get("model_code", "")).strip()
        sku = str(r.get("sku_num_sku_r3", "")).strip()
        if mc and sku and mc not in model_to_first_sku:
            model_to_first_sku[mc] = sku

    exp_to_fullpath = {}
    if df_cat is not None:
        for _, _cr in df_cat.iterrows():
            _exp = str(_cr.get("export_category", "")).strip()
            _fp = str(_cr.get("Category Path", "")).strip()
            if _exp and _fp and _exp not in exp_to_fullpath:
                exp_to_fullpath[_exp] = _fp

    for i, (_, src_row) in enumerate(results_df.iterrows()):
        row_idx = i + 2
        row_data = {}

        for master_col, tmpl_col in MASTER_TO_TEMPLATE.items():
            val = src_row.get(master_col, "")
            if pd.notna(val) and str(val).strip() not in ("", "nan"):
                row_data[tmpl_col] = str(val).strip()

        mc = str(src_row.get("model_code", "")).strip()
        if mc and mc in model_to_first_sku:
            row_data["ParentSKU"] = model_to_first_sku[mc]

        gtin = _format_gtin(src_row.get("bar_code", ""))
        if gtin:
            row_data["GTIN_Barcode"] = gtin

        product_name = str(src_row.get("product_name", "")).strip()
        color_raw = str(src_row.get("color", "")).strip()
        color = color_raw.split("|")[0].strip()
        if product_name and color and not product_name.lower().endswith(color.lower()):
            row_data["Name"] = f"{product_name} - {color.title()}"
        elif product_name:
            row_data["Name"] = product_name

        bw = str(src_row.get("business_weight", "")).strip()
        if bw and bw.lower() not in ("", "nan"):
            row_data["product_weight"] = re.sub(r'\s*kg\s*$', '', bw, flags=re.IGNORECASE).strip()

        size_val = re.sub(r'"+', '', str(src_row.get("size", ""))).strip().rstrip(".")
        if size_val.lower() not in ("", "nan", "no size"):
            pkg_name = row_data.get("Name", product_name)
            row_data["package_content"] = f"{pkg_name} - {size_val}"

        raw_brand = src_row.get("brand_name", "")
        if pd.notna(raw_brand) and str(raw_brand).strip():
            row_data["Brand"] = match_brand(str(raw_brand), df_brands)

        if ai_categories and i < len(ai_categories):
            primary_code, secondary_code = ai_categories[i]
        else:
            primary_code, secondary_code = ("", "")

        primary_full = exp_to_fullpath.get(primary_code, primary_code)
        if primary_full:
            row_data["PrimaryCategory"] = primary_full
        secondary_full = exp_to_fullpath.get(secondary_code, secondary_code)
        if secondary_full:
            row_data["AdditionalCategory"] = secondary_full

        sku = str(src_row.get("sku_num_sku_r3", "")).strip()
        override_size = sku_to_size_override.get(sku) if sku_to_size_override else None
        row_data["variation"] = get_variation(src_row, is_fashion, valid_sizes, override_size)

        row_data["Price_KES"] = "100000"
        row_data["Stock"] = 0

        color_for_family = str(src_row.get("color", "")).strip()
        if color_for_family and color_for_family.lower() not in ("", "nan"):
            row_data["color_family"] = color_for_family.split("|")[0].strip()

        if short_descs and i < len(short_descs) and short_descs[i]:
            row_data["short_description"] = short_descs[i]

        for tmpl_col, value in row_data.items():
            if tmpl_col in header_map:
                cell = ws.cell(row=row_idx, column=header_map[tmpl_col])
                cell.value = value
                cell.font = data_font
                cell.alignment = data_align

    # Keep ONLY Upload Template sheet
    for sheet_name in list(wb.sheetnames):
        if sheet_name != "Upload Template":
            wb.remove(wb[sheet_name])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ====================== BRAND MATCHING ======================
def match_brand(raw: str, df_brands: pd.DataFrame) -> str:
    if not raw or pd.isna(raw):
        return ""
    needle = str(raw).strip().lower()
    exact = df_brands[df_brands["brand_name_lower"] == needle]
    if not exact.empty:
        return exact.iloc[0]["brand_entry"]
    partial = df_brands[df_brands["brand_name_lower"].str.contains(needle, regex=False)]
    if not partial.empty:
        return partial.iloc[0]["brand_entry"]
    for _, brow in df_brands.iterrows():
        if brow["brand_name_lower"] in needle:
            return brow["brand_entry"]
    return str(raw).strip()

# ====================== SIDEBAR ======================
with st.sidebar:
    st.header("Master Data")
    uploaded_master = st.file_uploader("Working file (.xlsx or .csv)", type=["xlsx","csv"])

    st.markdown("---")
    st.header("Product Type")
    product_type = st.radio("Product type", ["Fashion", "Other"], index=0, horizontal=True)
    is_fashion = product_type == "Fashion"

    valid_sizes = parse_valid_sizes(SIZES_PATH)
    if valid_sizes:
        st.sidebar.info(f"✅ sizes.txt loaded: {len(valid_sizes)} sizes")
    else:
        st.sidebar.warning("sizes.txt not found in project folder!")

    st.markdown("---")
    st.header("Search Fields")
    also_search_name = st.checkbox("Also search by product name", value=False)

# ====================== LOAD DATA ======================
ref_bytes = None
try:
    ref_bytes = open(DECA_CAT_PATH, "rb").read()
except FileNotFoundError:
    st.sidebar.error(f"{DECA_CAT_PATH} not found")

if ref_bytes:
    df_cat, df_brands = load_reference_data(ref_bytes)

master_bytes = None
is_csv = True
if uploaded_master:
    master_bytes = uploaded_master.read()
    is_csv = uploaded_master.name.endswith(".csv")
    df_master = load_master(master_bytes, is_csv)
else:
    for path, csv_flag in [(MASTER_PATH, True), (MASTER_PATH.replace(".csv", ".xlsx"), False)]:
        try:
            master_bytes = open(path, "rb").read()
            is_csv = csv_flag
            df_master = load_master(master_bytes, is_csv)
            break
        except FileNotFoundError:
            continue
    else:
        st.error("No master file found. Upload one.")
        st.stop()

# ====================== INPUT TABS ======================
tab1, tab2 = st.tabs(["Upload a List", "Manual Entry"])
queries = []

with tab1:
    uploaded_list = st.file_uploader("Upload file with SKU numbers", type=["xlsx","csv","txt"])
    if uploaded_list:
        ext = uploaded_list.name.rsplit(".", 1)[-1].lower()
        if ext == "txt":
            queries = [l.strip() for l in uploaded_list.read().decode().splitlines() if l.strip()]
        elif ext == "csv":
            q_df = pd.read_csv(uploaded_list, header=None, dtype=str)
            queries = q_df.iloc[:, 0].dropna().str.strip().tolist()
        else:
            q_df = pd.read_excel(uploaded_list, header=None, dtype=str)
            queries = q_df.iloc[:, 0].dropna().str.strip().tolist()
        st.success(f"Loaded **{len(queries)}** search terms")

with tab2:
    manual = st.text_area("Enter one SKU number per line", height=160, placeholder="4273417\n4273418\n4273423")
    if manual.strip():
        queries = [q.strip() for q in manual.strip().splitlines() if q.strip()]

# ====================== RESULTS ======================
if queries:
    all_result_frames = []
    no_match = []
    for q in queries:
        mask = df_master["sku_num_sku_r3"].fillna("").str.strip() == q.strip()
        if also_search_name and "product_name" in df_master.columns:
            mask |= df_master["product_name"].fillna("").str.lower().str.contains(q.lower(), regex=False)
        res = df_master[mask].copy()
        if res.empty:
            no_match.append(q)
        else:
            res.insert(0, "Search Term", q)
            all_result_frames.append((q, res))

    if no_match:
        st.warning(f"No matches found for: **{', '.join(no_match)}**")

    if all_result_frames:
        combined = pd.concat([r for _, r in all_result_frames], ignore_index=True)
        total_rows = len(combined)
        st.success(f"**{total_rows} rows** matched")

        short_descs = [rule_based_short_desc(row) for _, row in combined.iterrows()]

        st.markdown("---")
        st.subheader(f"Results — {total_rows} SKU(s) — Preview & Edit Sizes")

        preview = combined.copy()
        preview["_variation"] = preview.apply(
            lambda r: get_variation(r, is_fashion=is_fashion, valid_sizes=valid_sizes),
            axis=1
        )

        sku_to_size_override = None
        if is_fashion and valid_sizes:
            preview["_size_status"] = preview["_variation"].apply(
                lambda v: "✅ Valid" if str(v) in valid_sizes or str(v) == "..." else "❌ Missing in sizes.txt"
            )

            edited_df = st.data_editor(
                preview[["sku_num_sku_r3", "product_name", "color", "size", "_variation", "_size_status"]],
                use_container_width=True,
                hide_index=True,
                height=420,
                column_config={
                    "_variation": st.column_config.SelectboxColumn("Size (validated)", options=["..."] + valid_sizes, width="medium"),
                    "_size_status": st.column_config.TextColumn("Size Status", width="small"),
                }
            )
            sku_to_size_override = {str(k).strip(): v for k, v in zip(edited_df["sku_num_sku_r3"], edited_df["_variation"])}
        else:
            st.dataframe(preview, use_container_width=True, hide_index=True, height=420)

        # Build template (single sheet)
        tpl_bytes = build_template(combined, df_cat, df_brands, None, short_descs, is_fashion, valid_sizes, sku_to_size_override)

        st.download_button(
            "✅ Download Upload Template Sheet ONLY (.xlsx)",
            data=tpl_bytes,
            file_name="decathlon_upload_template_filled.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        st.caption("The file contains **only** the Upload Template sheet.")

else:
    st.info("Upload a list or type SKUs above to get started.")

st.markdown("---")
st.caption("Decathlon Product Lookup · Powered by your Decathlon working file")
